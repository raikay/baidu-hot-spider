import requests
from bs4 import BeautifulSoup
import time
import os
import json
from datetime import datetime
import openpyxl

def fetch_baidu_hot():
    """爬取百度热搜榜数据"""
    url = "https://top.baidu.com/board?tab=realtime"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.encoding = 'utf-8'
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        results = []
        
        # 获取所有热搜卡片容器
        # 使用更精确的选择器来定位每个完整的热搜项目
        hot_list = soup.select('.category-wrap_iQLoo tbody')
        if hot_list:
            # 从tbody中获取所有行
            rows = hot_list[0].find_all('tr')[:20]  # 取前20个
            
            # 获取所有热搜指数元素
            hot_indices = soup.select('.hot-index_1Bl1a')
            
            for i, row in enumerate(rows, 1):
                # 从当前行获取标题
                title_element = row.select_one('.c-single-text-ellipsis')
                title = title_element.get_text().strip() if title_element else "无标题"
                
                # 从当前行获取简介
                desc_element = row.select_one('.hot-desc_1m_jR')
                description = desc_element.get_text().strip() if desc_element else "无简介"
                
                # 获取对应的热搜指数
                hot_index = "无指数"
                if i-1 < len(hot_indices):
                    hot_index = hot_indices[i-1].get_text().strip()
                
                results.append({
                    'rank': i,
                    'title': title,
                    'description': description,
                    'hot_index': hot_index
                })
        else:
            # 备用方法：使用内容卡片和指数分别获取
            # 先获取所有内容卡片
            content_cards = soup.select('.content_1YWBm')[:20]
            # 获取所有热搜指数
            hot_indices = soup.select('.hot-index_1Bl1a')[:20]
            
            # 确保数量匹配
            min_count = min(len(content_cards), len(hot_indices))
            
            for i in range(min_count):
                card = content_cards[i]
                # 从卡片中获取标题
                title_element = card.select_one('.c-single-text-ellipsis')
                title = title_element.get_text().strip() if title_element else "无标题"
                
                # 从卡片中获取简介
                desc_element = card.select_one('.hot-desc_1m_jR')
                description = desc_element.get_text().strip() if desc_element else "无简介"
                
                # 获取对应的热搜指数
                hot_index = hot_indices[i].get_text().strip()
                
                results.append({
                    'rank': i + 1,
                    'title': title,
                    'description': description,
                    'hot_index': hot_index
                })
        
        # 打印调试信息，确保简介不重复
        if len(results) > 1:
            print(f"数据检查: 第一条简介长度={len(results[0]['description'])}, 第二条简介长度={len(results[1]['description'])}")
            # 比较前两条简介是否相同
            if results[0]['description'] != results[1]['description']:
                print("✓ 简介匹配正常，无重复")
            else:
                print("! 警告：简介可能有重复")
        
        print(f"爬取完成，共获取 {len(results)} 条热搜数据")
        if results:
            print(f"样例数据: 标题='{results[0]['title']}', 简介前20字='{results[0]['description'][:20]}...', 热搜指数='{results[0]['hot_index']}'")
        
        return results
    except Exception as e:
        print(f"爬取失败: {e}")
        return []

def save_to_excel(data):
    """将爬取的数据转换为JSON并追加到同一个Excel文件中"""
    filename = "baidu_hot_history.xlsx"
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 将数据转换为JSON字符串
    json_data = json.dumps(data, ensure_ascii=False, indent=2)
    
    try:
        # 检查文件是否存在
        if os.path.exists(filename):
            # 打开现有文件
            workbook = openpyxl.load_workbook(filename)
            worksheet = workbook.active
        else:
            # 创建新文件和工作表
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "hot_search_history"
            # 设置表头
            worksheet.append(["爬取时间", "JSON数据"])
        
        # 追加新数据行
        worksheet.append([current_time, json_data])
        
        # 调整列宽
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 150
        
        # 保存文件
        workbook.save(filename)
        print(f"数据已追加到 {filename} 文件")
        return filename
    except Exception as e:
        print(f"保存数据失败: {e}")
        # 创建备份文件作为备选
        backup_filename = f"baidu_hot_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "hot_search_backup"
        worksheet.append(["爬取时间", "JSON数据"])
        worksheet.append([current_time, json_data[:32767]])  # 避免Excel单元格长度限制
        workbook.save(backup_filename)
        print(f"已创建备份文件: {backup_filename}")
        return backup_filename

def main():
    """主函数"""
    print(f"开始爬取百度热搜榜 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    data = fetch_baidu_hot()
    if data:
        save_to_excel(data)
    else:
        print("没有获取到数据")

if __name__ == "__main__":
    main()